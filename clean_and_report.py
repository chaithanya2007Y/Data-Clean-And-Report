"""
clean_and_report.py
--------------------
Data Cleaning & Reporting Automation

What it does:
 1. Loads a raw CSV dataset (raw_data.csv by default)
 2. Cleans it:
      - standardizes text (trims spaces, fixes case)
      - fixes inconsistent category values
      - parses/repairs date formats
      - handles missing values (fill / drop, column-specific rules)
      - removes duplicate rows
      - removes invalid values (e.g. negative units sold)
 3. Saves the cleaned dataset to cleaned_data.csv
 4. Generates an automated Excel report (report.xlsx) with:
      - a "Cleaned Data" sheet
      - a "Summary" sheet (KPIs + pivot-style aggregation)
      - a native Excel chart
 5. Generates a PNG visual summary (summary_chart.png)
 6. Prints a cleaning log to the console

Run:
    python clean_and_report.py
    python clean_and_report.py --input your_file.csv
"""

import argparse
import sys
from datetime import datetime

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")  # no GUI needed, safe for automation/scripts
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment


# --------------------------------------------------------------------------
# 1. LOAD
# --------------------------------------------------------------------------
def load_data(path: str) -> pd.DataFrame:
    print(f"[1/5] Loading data from '{path}' ...")
    df = pd.read_csv(path)
    print(f"      Loaded {len(df)} rows, {len(df.columns)} columns.")
    return df


# --------------------------------------------------------------------------
# 2. CLEAN
# --------------------------------------------------------------------------
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    print("[2/5] Cleaning data ...")
    log = []
    df = df.copy()

    # --- Standardize text columns (trim whitespace, consistent case) ---
    text_cols = ["Region", "Product", "SalesRep"]
    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.title()
            df[col] = df[col].replace({"None": np.nan, "Nan": np.nan})

    # --- Remove exact duplicate rows ---
    before = len(df)
    df = df.drop_duplicates()
    log.append(f"Removed {before - len(df)} duplicate rows.")

    # --- Parse inconsistent date formats ---
    if "OrderDate" in df.columns:
        df["OrderDate"] = pd.to_datetime(df["OrderDate"], errors="coerce", format="mixed")
        bad_dates = df["OrderDate"].isna().sum()
        log.append(f"Found {bad_dates} unparseable/invalid dates (set to NaT).")

    # --- Handle invalid numeric values (e.g. negative units) ---
    if "UnitsSold" in df.columns:
        invalid_units = (df["UnitsSold"] < 0).sum()
        df.loc[df["UnitsSold"] < 0, "UnitsSold"] = np.nan
        log.append(f"Found {invalid_units} negative UnitsSold values (set to missing).")

    # --- Handle missing values (column-specific strategy) ---
    missing_before = df.isna().sum().sum()

    # Numeric columns -> fill with median
    for col in ["UnitsSold", "UnitPrice"]:
        if col in df.columns:
            median_val = df[col].median()
            n_missing = df[col].isna().sum()
            df[col] = df[col].fillna(median_val)
            log.append(f"Filled {n_missing} missing '{col}' values with median ({median_val}).")

    # Categorical columns -> fill with 'Unknown'
    for col in text_cols:
        if col in df.columns:
            n_missing = df[col].isna().sum()
            df[col] = df[col].fillna("Unknown")
            log.append(f"Filled {n_missing} missing '{col}' values with 'Unknown'.")

    # Rows with missing critical OrderDate -> drop (can't report on them reliably)
    if "OrderDate" in df.columns:
        before_drop = len(df)
        df = df.dropna(subset=["OrderDate"])
        log.append(f"Dropped {before_drop - len(df)} rows with missing/invalid OrderDate.")

    missing_after = df.isna().sum().sum()
    log.append(f"Total missing values: {missing_before} -> {missing_after}")

    # --- Add a computed column useful for reporting ---
    if "UnitsSold" in df.columns and "UnitPrice" in df.columns:
        df["TotalSale"] = df["UnitsSold"] * df["UnitPrice"]

    print("      Cleaning steps:")
    for entry in log:
        print(f"        - {entry}")

    return df


# --------------------------------------------------------------------------
# 3. SAVE CLEANED DATA
# --------------------------------------------------------------------------
def save_cleaned_data(df: pd.DataFrame, path: str = "cleaned_data.csv"):
    print(f"[3/5] Saving cleaned data to '{path}' ...")
    df.to_csv(path, index=False)


# --------------------------------------------------------------------------
# 4. VISUAL SUMMARY (PNG)
# --------------------------------------------------------------------------
def create_visual_summary(df: pd.DataFrame, path: str = "summary_chart.png"):
    print(f"[4/5] Creating visual summary chart '{path}' ...")
    if "Region" not in df.columns or "TotalSale" not in df.columns:
        print("      Skipped (required columns not present).")
        return

    summary = df.groupby("Region")["TotalSale"].sum().sort_values(ascending=False)

    fig, ax = plt.subplots(figsize=(8, 5))
    summary.plot(kind="bar", ax=ax, color="#4472C4")
    ax.set_title("Total Sales by Region")
    ax.set_xlabel("Region")
    ax.set_ylabel("Total Sales ($)")
    plt.xticks(rotation=45)
    plt.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)


# --------------------------------------------------------------------------
# 5. EXCEL REPORT (with native chart + KPI summary)
# --------------------------------------------------------------------------
def create_excel_report(df: pd.DataFrame, path: str = "report.xlsx"):
    print(f"[5/5] Generating Excel report '{path}' ...")
    wb = Workbook()

    # ---- Sheet 1: Cleaned Data ----
    ws_data = wb.active
    ws_data.title = "Cleaned Data"
    for row in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(row)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws_data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col_cells in ws_data.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_data.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 30)

    # ---- Sheet 2: Summary ----
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Automated Data Report"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    ws_sum["A4"] = "KPI"
    ws_sum["B4"] = "Value"
    ws_sum["A4"].font = ws_sum["B4"].font = Font(bold=True)

    total_sales = df["TotalSale"].sum() if "TotalSale" in df.columns else 0
    total_orders = len(df)
    avg_order = df["TotalSale"].mean() if "TotalSale" in df.columns else 0

    kpis = [
        ("Total Orders", total_orders),
        ("Total Sales ($)", round(total_sales, 2)),
        ("Average Order Value ($)", round(avg_order, 2)),
    ]
    for i, (k, v) in enumerate(kpis, start=5):
        ws_sum[f"A{i}"] = k
        ws_sum[f"B{i}"] = v

    # Regional breakdown table (used to build the native chart)
    start_row = 10
    ws_sum[f"A{start_row}"] = "Region"
    ws_sum[f"B{start_row}"] = "Total Sales"
    ws_sum[f"A{start_row}"].font = ws_sum[f"B{start_row}"].font = Font(bold=True)

    if "Region" in df.columns and "TotalSale" in df.columns:
        region_summary = df.groupby("Region")["TotalSale"].sum().sort_values(ascending=False)
        for i, (region, total) in enumerate(region_summary.items(), start=start_row + 1):
            ws_sum[f"A{i}"] = region
            ws_sum[f"B{i}"] = round(total, 2)

        # Native Excel bar chart
        chart = BarChart()
        chart.title = "Total Sales by Region"
        chart.y_axis.title = "Sales ($)"
        chart.x_axis.title = "Region"
        data_ref = Reference(
            ws_sum, min_col=2, min_row=start_row, max_row=start_row + len(region_summary)
        )
        cats_ref = Reference(
            ws_sum, min_col=1, min_row=start_row + 1, max_row=start_row + len(region_summary)
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_sum.add_chart(chart, "D10")

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 18

    wb.save(path)


# --------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Data Cleaning & Reporting Automation")
    parser.add_argument("--input", default="raw_data.csv", help="Path to raw input CSV")
    parser.add_argument("--output", default="cleaned_data.csv", help="Path to save cleaned CSV")
    parser.add_argument("--report", default="report.xlsx", help="Path to save Excel report")
    args = parser.parse_args()

    try:
        df = load_data(args.input)
    except FileNotFoundError:
        print(f"ERROR: '{args.input}' not found. Run generate_sample_data.py first, "
              f"or pass --input your_file.csv")
        sys.exit(1)

    cleaned = clean_data(df)
    save_cleaned_data(cleaned, args.output)
    create_visual_summary(cleaned)
    create_excel_report(cleaned, args.report)

    print("\nDone! Files created:")
    print(f"  - {args.output}")
    print(f"  - summary_chart.png")
    print(f"  - {args.report}")


if __name__ == "__main__":
    main()
